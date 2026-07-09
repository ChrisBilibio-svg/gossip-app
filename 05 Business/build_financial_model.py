"""
Fofoca financial plan generator — v1.0
Builds a 36-month, multi-scenario workbook with two strategic funding paths
(bootstrapped vs funded), an explicit hiring ramp, a cap table with dilution,
and a shareholder-value (valuation) trajectory.

This is a PLANNING model built on explicit, defensible assumptions — not a
forecast or a promise. Replace assumptions with observed beta data when available.

Run:  python build_financial_model.py
Out:  fofoca_financial_model.xlsx  (in the same folder)
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

MONTHS = 36
USD = 5.30  # R$ per US$ (planning rate)

# ----------------------------------------------------------------------------
# 1. ASSUMPTIONS
# ----------------------------------------------------------------------------
A = {
    "start_mau": 1000,                # month-1 beta cohort, not a viral launch
    "dau_mau": 0.25,                  # daily/monthly active ratio
    "bets_per_dau_day": 3.0,          # engagement -> prediction volume
    # Monetization (introduced gradually as retention proves out)
    "ad_ecpm_brl": 6.0,               # R$ per 1000 impressions (thin early)
    "imps_per_dau_day": 6,            # ad impressions per active user/day
    "ad_start_month": 6,
    "pro_price_brl": 19.90,           # monthly Pro subscription
    "pro_conv_base": 0.018,           # % of MAU on Pro at maturity (base)
    "pro_start_month": 9,
    "pulse_start_month": 15,          # B2B Gossip Pulse dashboards
}

# Scenario growth curves: month-over-month MAU growth, tapering over time.
# Conservative assumes a fragile cold-start; Base a healthy single-wedge habit;
# Upside a wedge that compounds into adjacent fandoms. All start from a 1k beta.
SCENARIOS = {
    "Conservative": {"g0": 0.16, "g_decay": 0.92, "g_floor": 0.010, "pro_mult": 0.7, "pulse_step": 3500},
    "Base":         {"g0": 0.32, "g_decay": 0.95, "g_floor": 0.025, "pro_mult": 1.0, "pulse_step": 7000},
    "Upside":       {"g0": 0.40, "g_decay": 0.96, "g_floor": 0.045, "pro_mult": 1.3, "pulse_step": 12000},
}


def project(scen):
    """Return list of monthly dicts for a given scenario's top-line + revenue."""
    p = SCENARIOS[scen]
    rows = []
    mau = A["start_mau"]
    g = p["g0"]
    for m in range(1, MONTHS + 1):
        if m > 1:
            mau = mau * (1 + g)
            g = max(g * p["g_decay"], p["g_floor"])
        dau = mau * A["dau_mau"]
        bets = dau * A["bets_per_dau_day"] * 30

        ad_rev = 0.0
        if m >= A["ad_start_month"]:
            imps = dau * A["imps_per_dau_day"] * 30
            ad_rev = imps / 1000 * A["ad_ecpm_brl"]

        pro_rev = 0.0
        if m >= A["pro_start_month"]:
            ramp = min(1.0, (m - A["pro_start_month"] + 1) / 9.0)  # ramp conversion over ~9 mo
            conv = A["pro_conv_base"] * p["pro_mult"] * ramp
            pro_rev = mau * conv * A["pro_price_brl"]

        pulse_rev = 0.0
        if m >= A["pulse_start_month"]:
            # step up roughly every 6 months as B2B logos are added
            steps = 1 + (m - A["pulse_start_month"]) // 6
            pulse_rev = p["pulse_step"] * steps

        total_rev = ad_rev + pro_rev + pulse_rev
        rows.append({
            "month": m, "mau": mau, "dau": dau, "bets": bets,
            "ad": ad_rev, "pro": pro_rev, "pulse": pulse_rev,
            "rev": total_rev, "arpu": (total_rev / mau if mau else 0),
        })
    return rows


# ----------------------------------------------------------------------------
# 2. HIRING PLANS (loaded monthly cost in R$) — the investment/headcount ramp
# ----------------------------------------------------------------------------
# Each role: (title, loaded_monthly_cost, start_month_bootstrapped, start_month_funded)
# start_month = 0 means "not hired on this path".
ROLES = [
    ("Founder / CEO (Chris)",            6000,  1,  1),
    ("Curation Lead",                    8000,  3,  2),
    ("Mobile / Full-stack Dev",         14000,  7,  4),
    ("Backend / Data Engineer",         14000,  0,  7),   # funded only
    ("Growth / Marketing Lead",         12000, 14,  6),
    ("Community & Moderation",           5000, 10,  6),
    ("Community & Moderation #2",        5000,  0, 13),   # funded only
    ("BD / Sales (Pulse B2B)",          10000, 20, 14),
    ("Product Designer (contract)",      8000,  0,  9),   # funded only
    ("Legal / Trust (fractional)",       6000,  6,  4),
]


def headcount_cost(path, m):
    """Total monthly people cost for a path at month m (path: 'boot' or 'fund')."""
    idx = 2 if path == "boot" else 3
    total = 0
    heads = 0
    for r in ROLES:
        start = r[idx]
        if start and m >= start:
            total += r[1]
            heads += 1
    return total, heads


def nonpeople_opex(path, m, mau, bets):
    """Infra, content/AI, ads/marketing, tools, legal one-offs. Scales with usage."""
    # Infra + AI summarization scale with volume; marketing differs by path.
    infra = 800 + bets / 1000 * 1.2          # supabase + hosting + AI haiku calls
    tools = 1500
    if path == "boot":
        marketing = 1500 + mau * 0.05         # lean, organic-led
    else:
        marketing = 6000 + mau * 0.18         # paid acceleration after seed
    return infra + tools + marketing


# ----------------------------------------------------------------------------
# 3. FUNDING ROUNDS + CAP TABLE (funded path)
# ----------------------------------------------------------------------------
# Founders start: Chris 70%, Co-dev (Pedro) 20%, Option pool 10%.
CAP_START = {"Chris (Founder)": 0.70, "Co-dev / Pedro": 0.20, "Option Pool": 0.10}

ROUNDS = [
    # (name, month, raise_brl, pre_money_brl)
    ("Seed",     7,  1_500_000,  6_000_000),
    ("Series A", 24, 8_000_000, 32_000_000),
]

VAL_MULTIPLE = 6.0  # enterprise value = 6x ARR (consumer+B2B blended planning multiple)


def cap_table_after_rounds():
    """Return list of (round_label, table_dict, post_money) showing dilution."""
    table = dict(CAP_START)
    history = [("Founding", dict(table), None)]
    for name, month, raise_brl, pre in ROUNDS:
        post = pre + raise_brl
        new_investor_pct = raise_brl / post
        factor = 1 - new_investor_pct
        for k in table:
            table[k] *= factor
        table[f"{name} Investors"] = new_investor_pct
        history.append((name, dict(table), post))
    return history


# ----------------------------------------------------------------------------
# 4. BUILD P&L for a path (bootstrapped or funded), base scenario by default
# ----------------------------------------------------------------------------
def pnl(path, scen="Base"):
    proj = project(scen)
    cash_injections = {}
    if path == "fund":
        for name, month, raise_brl, pre in ROUNDS:
            cash_injections[month] = cash_injections.get(month, 0) + raise_brl
    rows = []
    cum = 0.0
    for r in proj:
        m = r["month"]
        people, heads = headcount_cost(path, m)
        other = nonpeople_opex(path, m, r["mau"], r["bets"])
        opex = people + other
        net = r["rev"] - opex
        inject = cash_injections.get(m, 0)
        cum += net + inject
        rows.append({**r, "people": people, "heads": heads, "other": other,
                     "opex": opex, "net": net, "inject": inject, "cum_cash": cum})
    return rows


# ----------------------------------------------------------------------------
# 5. EXCEL STYLING HELPERS
# ----------------------------------------------------------------------------
PINK = "FF4D9D"
DARK = "0F1419"
LIGHT = "FFF3F8"
GREY = "536471"

hdr_fill = PatternFill("solid", fgColor=PINK)
sub_fill = PatternFill("solid", fgColor=LIGHT)
hdr_font = Font(bold=True, color="FFFFFF", size=11)
title_font = Font(bold=True, color=DARK, size=16)
sub_font = Font(bold=True, color=DARK, size=11)
muted_font = Font(color=GREY, size=10)
thin = Side(style="thin", color="EFF3F4")
border = Border(bottom=thin)


def style_header(ws, row, ncols, start=1):
    for c in range(start, start + ncols):
        cell = ws.cell(row=row, column=c)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def money(cell):
    cell.number_format = u'R$ #,##0'


def pct(cell):
    cell.number_format = '0.0%'


def autosize(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ----------------------------------------------------------------------------
# 6. WRITE WORKBOOK
# ----------------------------------------------------------------------------
wb = Workbook()

# --- Cover ---
ws = wb.active
ws.title = "Cover"
ws["A1"] = "Fofoca — Financial Plan v1.0"
ws["A1"].font = title_font
ws["A2"] = "Brazil-first gossip prediction game · 36-month planning model"
ws["A2"].font = muted_font
notes = [
    "",
    "PLANNING MODEL — NOT A FORECAST OR GUARANTEE.",
    "Built on explicit, defensible assumptions (see Assumptions tab). Replace with",
    "observed beta data once available. All figures in BRL unless noted. US$ at R$ %.2f." % USD,
    "",
    "Tabs:",
    "  • Assumptions        — every editable driver behind the model",
    "  • Scenarios          — Conservative / Base / Upside top-line + revenue (36 mo)",
    "  • Bootstrapped P&L   — profit-funded hiring, founders keep 100%",
    "  • Funded P&L         — seed + Series A capital, faster hiring ramp",
    "  • Hiring Plan        — role-by-role investment ramp per path",
    "  • Cap Table          — ownership & dilution across rounds (funded path)",
    "  • Shareholder Value  — valuation trajectory & founder equity value",
    "  • Comparison         — bootstrapped vs funded, side by side",
    "",
    "Revenue thesis: free core prediction game (habit + data) → light ads → Pro",
    "subscription (speed/status, never pay-to-win) → Gossip Pulse B2B trend data.",
    "Money never touches truth or scoring.",
]
for i, line in enumerate(notes, 4):
    ws.cell(row=i, column=1, value=line)
    if "NOT A FORECAST" in line:
        ws.cell(row=i, column=1).font = Font(bold=True, color="B00020")
autosize(ws, [95])

# --- Assumptions ---
ws = wb.create_sheet("Assumptions")
ws["A1"] = "Assumptions (editable drivers)"
ws["A1"].font = title_font
rows = [
    ("Driver", "Value", "Note"),
    ("Start MAU (month 1)", A["start_mau"], "Private beta cohort, not viral launch"),
    ("DAU / MAU ratio", A["dau_mau"], "Daily engagement"),
    ("Bets per DAU per day", A["bets_per_dau_day"], "Prediction volume"),
    ("Ad eCPM (R$)", A["ad_ecpm_brl"], "Thin early; secondary line"),
    ("Ad impressions / DAU / day", A["imps_per_dau_day"], ""),
    ("Ads start (month)", A["ad_start_month"], "Only after retention proven"),
    ("Pro price (R$/mo)", A["pro_price_brl"], "Speed/status, never pay-to-win"),
    ("Pro conversion at maturity (Base)", A["pro_conv_base"], "% of MAU; ramps over 9 mo"),
    ("Pro start (month)", A["pro_start_month"], ""),
    ("Pulse B2B start (month)", A["pulse_start_month"], "Needs data density first"),
    ("Valuation multiple (x ARR)", VAL_MULTIPLE, "Blended consumer+B2B planning multiple"),
    ("FX rate (R$/US$)", USD, "Planning rate"),
]
for i, (a, b, c) in enumerate(rows, 3):
    ws.cell(row=i, column=1, value=a)
    ws.cell(row=i, column=2, value=b)
    ws.cell(row=i, column=3, value=c)
    if i == 3:
        style_header(ws, i, 3)
    else:
        ws.cell(row=i, column=1).font = sub_font
        ws.cell(row=i, column=3).font = muted_font
autosize(ws, [34, 16, 46])

# --- Scenarios (revenue top-line) ---
ws = wb.create_sheet("Scenarios")
ws["A1"] = "Scenarios — top-line & revenue (36 months)"
ws["A1"].font = title_font
headers = ["Scenario", "Month", "MAU", "DAU", "Bets/mo", "Ad R$", "Pro R$", "Pulse R$", "Total Rev R$", "ARPU R$"]
r0 = 3
for j, h in enumerate(headers, 1):
    ws.cell(row=r0, column=j, value=h)
style_header(ws, r0, len(headers))
rr = r0 + 1
for scen in SCENARIOS:
    proj = project(scen)
    for r in proj:
        if r["month"] in (1, 3, 6, 9, 12, 18, 24, 30, 36):  # milestone months for readability
            ws.cell(row=rr, column=1, value=scen)
            ws.cell(row=rr, column=2, value=r["month"])
            ws.cell(row=rr, column=3, value=round(r["mau"]))
            ws.cell(row=rr, column=4, value=round(r["dau"]))
            ws.cell(row=rr, column=5, value=round(r["bets"]))
            for col, key in zip((6, 7, 8, 9), ("ad", "pro", "pulse", "rev")):
                cell = ws.cell(row=rr, column=col, value=round(r[key]))
                money(cell)
            ws.cell(row=rr, column=10, value=round(r["arpu"], 2))
            rr += 1
    rr += 1  # blank line between scenarios
autosize(ws, [14, 8, 11, 10, 12, 11, 12, 12, 14, 10])

# --- P&L tabs (Bootstrapped + Funded), Base scenario ---
def write_pnl(sheet_name, path, title, scen):
    ws = wb.create_sheet(sheet_name)
    ws["A1"] = title
    ws["A1"].font = title_font
    ws["A2"] = f"{scen} scenario · monthly · R$"
    ws["A2"].font = muted_font
    headers = ["Month", "MAU", "Revenue", "Headcount", "People cost", "Other opex",
               "Total opex", "Net", "Capital in", "Cumulative cash"]
    r0 = 4
    for j, h in enumerate(headers, 1):
        ws.cell(row=r0, column=j, value=h)
    style_header(ws, r0, len(headers))
    data = pnl(path, scen)
    rr = r0 + 1
    for d in data:
        ws.cell(row=rr, column=1, value=d["month"])
        ws.cell(row=rr, column=2, value=round(d["mau"]))
        money(ws.cell(row=rr, column=3, value=round(d["rev"])))
        ws.cell(row=rr, column=4, value=d["heads"])
        money(ws.cell(row=rr, column=5, value=round(d["people"])))
        money(ws.cell(row=rr, column=6, value=round(d["other"])))
        money(ws.cell(row=rr, column=7, value=round(d["opex"])))
        money(ws.cell(row=rr, column=8, value=round(d["net"])))
        money(ws.cell(row=rr, column=9, value=round(d["inject"])))
        money(ws.cell(row=rr, column=10, value=round(d["cum_cash"])))
        rr += 1
    autosize(ws, [8, 11, 12, 11, 12, 12, 12, 12, 12, 16])
    return data

# Bootstrapped = organic Base growth; Funded = capital-accelerated Upside growth.
boot_data = write_pnl("Bootstrapped P&L", "boot", "Bootstrapped path — profit-funded hiring, founders keep 100% (organic/Base growth)", "Base")
fund_data = write_pnl("Funded P&L", "fund", "Funded path — seed + Series A, accelerated hiring (capital-driven Upside growth)", "Upside")

# --- Hiring Plan ---
ws = wb.create_sheet("Hiring Plan")
ws["A1"] = "Hiring & investment ramp"
ws["A1"].font = title_font
ws["A2"] = "Loaded monthly cost (R$). 'Start month' = when the role is added; '—' = not on that path."
ws["A2"].font = muted_font
headers = ["Role", "Loaded R$/mo", "Bootstrapped start", "Funded start"]
r0 = 4
for j, h in enumerate(headers, 1):
    ws.cell(row=r0, column=j, value=h)
style_header(ws, r0, len(headers))
rr = r0 + 1
for title, cost, b_start, f_start in ROLES:
    ws.cell(row=rr, column=1, value=title)
    money(ws.cell(row=rr, column=2, value=cost))
    ws.cell(row=rr, column=3, value=(f"M{b_start}" if b_start else "—"))
    ws.cell(row=rr, column=4, value=(f"M{f_start}" if f_start else "—"))
    rr += 1
# totals at M36
b_tot, b_heads = headcount_cost("boot", 36)
f_tot, f_heads = headcount_cost("fund", 36)
rr += 1
ws.cell(row=rr, column=1, value="TEAM AT MONTH 36").font = sub_font
ws.cell(row=rr, column=3, value=f"{b_heads} people · R$ {b_tot:,.0f}/mo")
ws.cell(row=rr, column=4, value=f"{f_heads} people · R$ {f_tot:,.0f}/mo")
autosize(ws, [30, 14, 20, 16])

# --- Cap Table ---
ws = wb.create_sheet("Cap Table")
ws["A1"] = "Cap table & dilution (funded path)"
ws["A1"].font = title_font
ws["A2"] = "Ownership % after each round. Founders start 100%; capital dilutes everyone pro-rata."
ws["A2"].font = muted_font
history = cap_table_after_rounds()
holders = list(history[-1][1].keys())  # full set including round investors
r0 = 4
ws.cell(row=r0, column=1, value="Holder")
for j, (label, table, post) in enumerate(history, 2):
    ws.cell(row=r0, column=j, value=label)
style_header(ws, r0, len(history) + 1)
for i, holder in enumerate(holders, r0 + 1):
    ws.cell(row=i, column=1, value=holder)
    for j, (label, table, post) in enumerate(history, 2):
        v = table.get(holder)
        if v is not None:
            pct(ws.cell(row=i, column=j, value=v))
# post-money row
pm_row = r0 + len(holders) + 2
ws.cell(row=pm_row, column=1, value="Post-money valuation").font = sub_font
for j, (label, table, post) in enumerate(history, 2):
    if post:
        money(ws.cell(row=pm_row, column=j, value=post))
# round detail
dr = pm_row + 2
ws.cell(row=dr, column=1, value="Rounds:").font = sub_font
for k, (name, month, raise_brl, pre) in enumerate(ROUNDS, 1):
    ws.cell(row=dr + k, column=1, value=f"{name} — Month {month}: raise R$ {raise_brl:,.0f} at R$ {pre:,.0f} pre-money")
autosize(ws, [22, 14, 14, 14])

# --- Shareholder Value ---
ws = wb.create_sheet("Shareholder Value")
ws["A1"] = "Shareholder value (valuation trajectory)"
ws["A1"].font = title_font
ws["A2"] = f"Enterprise value = {VAL_MULTIPLE:.0f}x ARR (ARR = month revenue x 12). Illustrative, not a promise."
ws["A2"].font = muted_font
headers = ["Scenario / Path", "M36 ARR R$", "Enterprise value R$", "Founders' stake", "Founders' value R$", "in US$"]
r0 = 4
for j, h in enumerate(headers, 1):
    ws.cell(row=r0, column=j, value=h)
style_header(ws, r0, len(headers))
final_cap = history[-1][1]
founder_pct_funded = final_cap["Chris (Founder)"] + final_cap["Co-dev / Pedro"]
rr = r0 + 1
for scen in SCENARIOS:
    proj = project(scen)
    arr = proj[-1]["rev"] * 12
    ev = arr * VAL_MULTIPLE
    # Bootstrapped: founders own 100% (minus option pool 10%) => 90%
    for path_label, fpct in (("Bootstrapped", 0.90), ("Funded", founder_pct_funded)):
        ws.cell(row=rr, column=1, value=f"{scen} · {path_label}")
        money(ws.cell(row=rr, column=2, value=round(arr)))
        money(ws.cell(row=rr, column=3, value=round(ev)))
        pct(ws.cell(row=rr, column=4, value=fpct))
        fv = ev * fpct
        money(ws.cell(row=rr, column=5, value=round(fv)))
        ws.cell(row=rr, column=6, value=f"US$ {fv/USD:,.0f}")
        rr += 1
    rr += 1
ws.cell(row=rr, column=1, value="Note: 'Funded' founder stake is post Seed + Series A dilution; capital trades ownership for growth & speed.").font = muted_font
autosize(ws, [22, 14, 18, 13, 16, 16])

# --- Comparison ---
ws = wb.create_sheet("Comparison")
ws["A1"] = "Bootstrapped vs Funded — side by side"
ws["A1"].font = title_font
ws["A2"] = "Bootstrapped = organic/Base growth · Funded = capital-accelerated/Upside growth. The trade: control & capital efficiency vs. speed & scale."
ws["A2"].font = muted_font

def path_summary(data, path):
    m12 = data[11]; m36 = data[35]
    peak_burn = min(d["cum_cash"] for d in data)  # most negative cumulative cash
    total_capital = sum(d["inject"] for d in data)
    return {
        "m12_mau": round(data[11]["mau"]),
        "m36_mau": round(m36["mau"]),
        "m36_rev": round(m36["rev"]),
        "m36_net": round(m36["net"]),
        "m36_heads": m36["heads"],
        "peak_cash_need": round(-peak_burn) if peak_burn < 0 else 0,
        "capital_raised": round(total_capital),
        "m36_cum_cash": round(m36["cum_cash"]),
    }

bs = path_summary(boot_data, "boot")
fs = path_summary(fund_data, "fund")
metrics = [
    ("Metric", "Bootstrapped", "Funded"),
    ("MAU at Month 12", f"{bs['m12_mau']:,}", f"{fs['m12_mau']:,}"),
    ("MAU at Month 36", f"{bs['m36_mau']:,}", f"{fs['m36_mau']:,}"),
    ("Revenue at Month 36 (R$/mo)", f"R$ {bs['m36_rev']:,}", f"R$ {fs['m36_rev']:,}"),
    ("Net at Month 36 (R$/mo)", f"R$ {bs['m36_net']:,}", f"R$ {fs['m36_net']:,}"),
    ("Team size at Month 36", f"{bs['m36_heads']}", f"{fs['m36_heads']}"),
    ("External capital raised", f"R$ {bs['capital_raised']:,}", f"R$ {fs['capital_raised']:,}"),
    ("Peak cash need (pre-funding)", f"R$ {bs['peak_cash_need']:,}", f"R$ {fs['peak_cash_need']:,}"),
    ("Cumulative cash at M36", f"R$ {bs['m36_cum_cash']:,}", f"R$ {fs['m36_cum_cash']:,}"),
    ("Founder ownership at M36", "~90% (option pool only)", f"~{founder_pct_funded*100:.0f}% (post Seed + A)"),
    ("Control", "Full founder control", "Shared w/ investors + board"),
    ("Risk profile", "Slower, capital-efficient, fragile to cold-start", "Faster scale, dilutive, higher burn"),
]
r0 = 4
for i, (a, b, c) in enumerate(metrics, r0):
    ws.cell(row=i, column=1, value=a)
    ws.cell(row=i, column=2, value=b)
    ws.cell(row=i, column=3, value=c)
    if i == r0:
        style_header(ws, i, 3)
    else:
        ws.cell(row=i, column=1).font = sub_font
autosize(ws, [34, 32, 32])

import os
out = os.path.join(os.path.dirname(os.path.abspath(__file__)), "FOFOCA_FINANCIAL_PLAN.xlsx")
wb.save(out)
print("Saved:", out)

# Print a quick console summary for the markdown writeup
print("\n--- SUMMARY ---")
for scen in SCENARIOS:
    proj = project(scen)
    m36 = proj[-1]
    print(f"{scen:12} M36 MAU={round(m36['mau']):>8,}  rev/mo=R${round(m36['rev']):>9,}  ARR=R${round(m36['rev']*12):>10,}")
print(f"\nBootstrapped: M36 MAU={bs['m36_mau']:,} rev=R${bs['m36_rev']:,}/mo net=R${bs['m36_net']:,}/mo peak cash need=R${bs['peak_cash_need']:,} team={bs['m36_heads']}")
print(f"Funded:       M36 MAU={fs['m36_mau']:,} rev=R${fs['m36_rev']:,}/mo net=R${fs['m36_net']:,}/mo raised=R${fs['capital_raised']:,} team={fs['m36_heads']}")
print(f"Founder ownership funded path after Seed+A: {founder_pct_funded*100:.1f}%")
